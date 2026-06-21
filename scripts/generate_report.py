"""
generate_report.py
Takes the KPI report dict (from compute_kpis.py) and generates
a formatted Excel report with a summary sheet and a regional breakdown.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from compute_kpis import main as compute_kpis_main


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")


def arrow(pct):
    if pct is None:
        return "N/A"
    arrow_symbol = "▲" if pct >= 0 else "▼"
    return f"{arrow_symbol} {pct}%"


def generate_excel_report(report, output_dir="../reports"):
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Summary"

    this_week = report["this_week_kpis"]
    last_week = report["last_week_kpis"]
    wow = report["wow_change"]
    window = report["window"]

    # --- Title ---
    ws.merge_cells("A1:D1")
    ws["A1"] = "Weekly KPI Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"This week: {window['this_week'][0]} to {window['this_week'][1]}"
    ws["A3"] = f"Last week: {window['last_week'][0]} to {window['last_week'][1]}"

    # --- KPI Table Header ---
    headers = ["Metric", "This Week", "Last Week", "WoW Change"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        style_header(cell)

    # --- KPI Rows ---
    rows = [
        ("Total Revenue", f"${this_week['total_revenue']:,.2f}",
         f"${last_week['total_revenue']:,.2f}", arrow(wow["revenue_pct"])),
        ("Total Orders", this_week["total_orders"],
         last_week["total_orders"], arrow(wow["orders_pct"])),
        ("Avg Order Value", f"${this_week['avg_order_value']:,.2f}",
         f"${last_week['avg_order_value']:,.2f}", arrow(wow["avg_order_value_pct"])),
        ("Top Category", this_week["top_category"], last_week["top_category"], ""),
    ]

    for i, row in enumerate(rows, start=6):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            if col == 4 and isinstance(value, str) and "▲" in value:
                cell.font = Font(color="006100")  # green for growth
            elif col == 4 and isinstance(value, str) and "▼" in value:
                cell.font = Font(color="9C0006")  # red for decline

    # --- Regional Breakdown Sheet ---
    ws2 = wb.create_sheet("Revenue by Region")
    ws2.cell(row=1, column=1, value="Region")
    ws2.cell(row=1, column=2, value="Revenue")
    style_header(ws2.cell(row=1, column=1))
    style_header(ws2.cell(row=1, column=2))

    for i, (region, revenue) in enumerate(this_week["revenue_by_region"].items(), start=2):
        ws2.cell(row=i, column=1, value=region)
        ws2.cell(row=i, column=2, value=revenue)

    # --- Auto-width columns ---
    for sheet in [ws, ws2]:
        for col_cells in sheet.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    # --- Save with timestamp ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"weekly_kpi_report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    print(f"✅ Excel report saved: {filepath}")
    return filepath


if __name__ == "__main__":
    report = compute_kpis_main()
    generate_excel_report(report)





   